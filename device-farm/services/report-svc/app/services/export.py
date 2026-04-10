# Export Service for Device Farm
from datetime import datetime
from typing import Optional, List, Dict, Any
import io
import csv
import logging
from pydantic import BaseModel

logger = logging.getLogger(__name__)

# Optional imports for PDF and Excel generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed - PDF export will be limited")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel export will be limited")


class ExportFormat:
    """Supported export formats"""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"


class ExportService:
    """Service for exporting data to various formats"""

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str,
    ) -> bytes:
        """Export data to CSV format"""
        if not data:
            return b""

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue().encode('utf-8')

    def export_statistics_to_csv(
        self,
        device_usage: List[Dict[str, Any]],
        task_stats: Optional[Dict[str, Any]] = None,
        usage_trend: Optional[List[Dict[str, Any]]] = None,
    ) -> bytes:
        """Export statistics data to CSV format (multiple sections)"""
        output = io.StringIO()

        # Device Usage Section
        if device_usage:
            output.write("Device Usage Statistics\n")
            writer = csv.DictWriter(output, fieldnames=device_usage[0].keys())
            writer.writeheader()
            writer.writerows(device_usage)
            output.write("\n\n")

        # Task Stats Section
        if task_stats:
            output.write("Task Execution Statistics\n")
            writer = csv.DictWriter(output, fieldnames=task_stats.keys())
            writer.writerow(task_stats)
            output.write("\n\n")

        # Usage Trend Section
        if usage_trend:
            output.write("Usage Trend\n")
            writer = csv.DictWriter(output, fieldnames=usage_trend[0].keys())
            writer.writeheader()
            writer.writerows(usage_trend)

        return output.getvalue().encode('utf-8')

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        sheet_name: str = "Data",
        filename: str = "export.xlsx",
    ) -> bytes:
        """Export data to Excel format"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, returning empty Excel")
            return b""

        if not data:
            return b""

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        # Write headers
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, key in enumerate(headers, 1):
                value = row_data.get(key, "")
                # Handle datetime serialization
                if isinstance(value, datetime):
                    value = value.isoformat()
                sheet.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_execution_logs_to_excel(
        self,
        logs: List[Dict[str, Any]],
        summary: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        """Export execution logs to Excel with summary sheet"""
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available, returning empty Excel")
            return b""

        workbook = openpyxl.Workbook()

        # Summary sheet
        if summary:
            summary_sheet = workbook.active
            summary_sheet.title = "Summary"

            row = 1
            for key, value in summary.items():
                summary_sheet.cell(row=row, column=1, value=key).font = Font(bold=True)
                cell = summary_sheet.cell(row=row, column=2, value=str(value) if not isinstance(value, datetime) else value.isoformat())
                row += 1

        # Logs sheet
        if logs:
            logs_sheet = workbook.create_sheet(title="Execution Logs")

            headers = list(logs[0].keys())
            for col, header in enumerate(headers, 1):
                cell = logs_sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

            for row_idx, log in enumerate(logs, 2):
                for col_idx, key in enumerate(headers, 1):
                    value = log.get(key, "")
                    if isinstance(value, datetime):
                        value = value.isoformat()
                    logs_sheet.cell(row=row_idx, column=col_idx, value=value)

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def export_to_pdf(
        self,
        title: str,
        content: Dict[str, Any],
        filename: str = "report.pdf",
    ) -> bytes:
        """Export data to PDF format"""
        if not REPORTLAB_AVAILABLE:
            logger.warning("reportlab not available, returning empty PDF")
            return b""

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 20))

        # Summary section
        if 'summary' in content:
            elements.append(Paragraph("Summary", styles['Heading2']))
            summary = content['summary']

            summary_data = [[k, str(v) if not isinstance(v, datetime) else v.isoformat()]
                           for k, v in summary.items()]
            summary_table = Table(summary_data, colWidths=[200, 300])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightblue),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))

        # Test cases section
        if 'test_cases' in content:
            elements.append(Paragraph("Test Cases", styles['Heading2']))
            test_cases = content['test_cases']

            # Header
            tc_data = [['Name', 'Status', 'Duration', 'Message']]
            for tc in test_cases[:50]:  # Limit to 50 for PDF
                tc_data.append([
                    str(tc.get('name', ''))[:40],
                    str(tc.get('status', '')),
                    str(tc.get('duration', '')),
                    str(tc.get('message', ''))[:50] if tc.get('message') else ''
                ])

            tc_table = Table(tc_data, colWidths=[150, 60, 60, 170])
            tc_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            elements.append(tc_table)

        # Device usage section
        if 'device_usage' in content:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("Device Usage", styles['Heading2']))
            device_usage = content['device_usage']

            du_data = [['Device', 'Usage (min)', 'Sessions', 'Last Used']]
            for du in device_usage[:20]:
                du_data.append([
                    str(du.get('device_name', du.get('device_id', '')))[:30],
                    str(round(du.get('total_usage_minutes', 0), 1)),
                    str(du.get('session_count', 0)),
                    str(du.get('last_used', ''))[:19] if du.get('last_used') else 'N/A'
                ])

            du_table = Table(du_data, colWidths=[150, 80, 80, 100])
            du_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 6),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
            ]))
            elements.append(du_table)

        doc.build(elements)
        return output.getvalue()

    def export_test_report_to_pdf(
        self,
        report: Dict[str, Any],
    ) -> bytes:
        """Export test report to PDF"""
        return self.export_to_pdf(
            title=f"Test Report - {report.get('title', 'Untitled')}",
            content={
                'summary': report.get('summary', {}),
                'test_cases': report.get('test_cases', []),
            }
        )

    def batch_export(
        self,
        exports: List[Dict[str, Any]],
    ) -> Dict[str, bytes]:
        """Perform batch export of multiple data sets"""
        results = {}

        for export_item in exports:
            export_type = export_item.get('type')
            data = export_item.get('data', [])
            filename = export_item.get('filename', 'export')
            format_type = export_item.get('format', 'csv')

            try:
                if format_type == ExportFormat.CSV:
                    if export_type == 'statistics':
                        results[filename] = self.export_statistics_to_csv(
                            device_usage=data.get('device_usage', []),
                            task_stats=data.get('task_stats'),
                            usage_trend=data.get('usage_trend'),
                        )
                    else:
                        results[filename] = self.export_to_csv(data, filename)

                elif format_type == ExportFormat.EXCEL:
                    if export_type == 'execution_logs':
                        results[filename] = self.export_execution_logs_to_excel(
                            logs=data.get('logs', []),
                            summary=data.get('summary'),
                        )
                    else:
                        results[filename] = self.export_to_excel(
                            data=data,
                            sheet_name=export_item.get('sheet_name', 'Data'),
                            filename=filename,
                        )

                elif format_type == ExportFormat.PDF:
                    results[filename] = self.export_to_pdf(
                        title=export_item.get('title', 'Report'),
                        content=data,
                        filename=filename,
                    )

            except Exception as e:
                logger.error(f"Failed to export {filename}: {e}")
                results[filename] = b""

        return results


# Global instance
export_service = ExportService()
